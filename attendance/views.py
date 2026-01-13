from django.shortcuts import render
from django.views import View
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils import timezone
from datetime import time, datetime
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.db.models import Count, Sum, Q

from .models import RegistroAsistencia
from employees.models import Empleado


class MarcarAsistenciaView(LoginRequiredMixin, View):
    """
    Vista para marcar asistencia vía AJAX.
    """

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            tipo = data.get('tipo', 'ENTRADA')
            latitud = data.get('latitud')
            longitud = data.get('longitud')
            empleado_id = data.get('empleado_id')

            # Obtener el empleado
            if empleado_id:
                # Si se proporciona empleado_id, marcar para ese empleado (requiere permisos de staff)
                if not request.user.is_staff:
                    return JsonResponse({
                        'success': False,
                        'message': 'No tienes permisos para marcar asistencia de otros empleados.'
                    }, status=403)
                try:
                    empleado = Empleado.objects.get(pk=empleado_id)
                except Empleado.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'message': 'Empleado no encontrado.'
                    }, status=404)
            else:
                # Si no se proporciona empleado_id, marcar para el usuario actual
                try:
                    empleado = request.user.empleado
                except AttributeError:
                    return JsonResponse({
                        'success': False,
                        'message': 'No tienes un perfil de empleado asociado.'
                    }, status=400)

            # Verificar si ya marcó entrada/salida hoy
            hoy = timezone.now().date()
            registros_hoy = RegistroAsistencia.objects.filter(
                empleado=empleado,
                fecha_hora__date=hoy
            )

            if tipo == 'ENTRADA':
                if registros_hoy.filter(tipo='ENTRADA').exists():
                    return JsonResponse({
                        'success': False,
                        'message': 'Ya has marcado entrada hoy.'
                    }, status=400)
            elif tipo == 'SALIDA':
                if not registros_hoy.filter(tipo='ENTRADA').exists():
                    return JsonResponse({
                        'success': False,
                        'message': 'Debes marcar entrada antes de marcar salida.'
                    }, status=400)
                if registros_hoy.filter(tipo='SALIDA').exists():
                    return JsonResponse({
                        'success': False,
                        'message': 'Ya has marcado salida hoy.'
                    }, status=400)

            # Calcular si es tardanza (para entrada)
            es_tardanza = False
            minutos_atraso = 0
            if tipo == 'ENTRADA':
                hora_limite = time(9, 0)  # 9:00 AM
                hora_actual = timezone.now().time()
                if hora_actual > hora_limite:
                    es_tardanza = True
                    # Calcular minutos de atraso
                    tiempo_limite = datetime.combine(timezone.now().date(), hora_limite)
                    tiempo_actual = datetime.combine(timezone.now().date(), hora_actual)
                    diferencia = tiempo_actual - tiempo_limite
                    minutos_atraso = int(diferencia.total_seconds() / 60)

            # Crear registro
            registro = RegistroAsistencia.objects.create(
                empleado=empleado,
                tipo=tipo,
                latitud=latitud,
                longitud=longitud,
                es_tardanza=es_tardanza,
                minutos_atraso=minutos_atraso
            )

            return JsonResponse({
                'success': True,
                'message': f'Asistencia marcada correctamente: {tipo}',
                'registro': {
                    'id': registro.id,
                    'tipo': registro.tipo,
                    'fecha_hora': registro.fecha_hora.strftime('%Y-%m-%d %H:%M:%S'),
                    'es_tardanza': registro.es_tardanza
                }
            })

        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Datos JSON inválidos.'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error interno del servidor: {str(e)}'
            }, status=500)


class ExportarAsistenciaExcelView(LoginRequiredMixin, View):
    """
    Vista para exportar reporte de asistencia en formato Excel para pre-nómina.
    """

    def get(self, request, *args, **kwargs):
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Pre-Nómina"

        # Estilos
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Encabezados
        headers = ['Nombre Empleado', 'Días Trabajados', 'Horas Extra (Calculadas)', 'Minutos de Atraso']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment

        # Obtener datos del mes actual
        hoy = timezone.now()
        inicio_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        fin_mes = (inicio_mes + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(days=1)

        # Obtener empleados activos
        empleados = Empleado.objects.filter(estado='activo')

        row_num = 2
        for empleado in empleados:
            # Calcular días trabajados (registros de entrada en el mes)
            dias_trabajados = RegistroAsistencia.objects.filter(
                empleado=empleado,
                tipo='ENTRADA',
                fecha_hora__date__range=[inicio_mes.date(), fin_mes.date()]
            ).count()

            # Calcular horas extra (simplificado: si hay más de 8 horas por día)
            # Por ahora, un cálculo básico - esto se puede mejorar
            horas_extra = 0
            registros_mes = RegistroAsistencia.objects.filter(
                empleado=empleado,
                fecha_hora__date__range=[inicio_mes.date(), fin_mes.date()]
            ).order_by('fecha_hora')

            # Agrupar por día y calcular horas trabajadas
            dias_procesados = {}
            entrada_actual = None

            for registro in registros_mes:
                fecha = registro.fecha_hora.date()
                if registro.tipo == 'ENTRADA':
                    entrada_actual = registro.fecha_hora
                elif registro.tipo == 'SALIDA' and entrada_actual and entrada_actual.date() == fecha:
                    horas_trabajadas = (registro.fecha_hora - entrada_actual).total_seconds() / 3600
                    if horas_trabajadas > 8:
                        horas_extra += horas_trabajadas - 8
                    entrada_actual = None

            # Calcular minutos de atraso (registros marcados como tardanza)
            minutos_atraso = RegistroAsistencia.objects.filter(
                empleado=empleado,
                tipo='ENTRADA',
                es_tardanza=True,
                fecha_hora__date__range=[inicio_mes.date(), fin_mes.date()]
            ).aggregate(total_atraso=Sum('minutos_atraso'))['total_atraso'] or 0

            # Escribir fila
            ws.cell(row=row_num, column=1, value=str(empleado))
            ws.cell(row=row_num, column=2, value=dias_trabajados)
            ws.cell(row=row_num, column=3, value=round(horas_extra, 2))
            ws.cell(row=row_num, column=4, value=minutos_atraso)

            row_num += 1

        # Ajustar ancho de columnas
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 20

        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=pre-nomina_{hoy.strftime("%Y-%m")}.xlsx'

        # Guardar workbook en la respuesta
        wb.save(response)

        return response
