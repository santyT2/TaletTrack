# attendance/views.py - VERSIÓN API PURA (HEADLESS)
# ========================================================
# Solo APIView y ViewSets de Django REST Framework
# Sin render(), LoginRequiredMixin - SOLO JSON
# ========================================================

from math import radians, cos, sin, asin, sqrt

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, viewsets, filters
from rest_framework.permissions import AllowAny
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Sum
from datetime import time, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    RegistroAsistencia,
    Turno,
    Geocerca,
    ReglaAsistencia,
    EventoAsistencia,
    JornadaCalculada,
)
from .serializers import (
    RegistroAsistenciaSerializer,
    TurnoSerializer,
    GeocercaSerializer,
    ReglaAsistenciaSerializer,
    EventoAsistenciaSerializer,
    JornadaCalculadaSerializer,
)
from employees.models import Empleado


class MarcarAsistenciaView(APIView):
    """
    API para marcar entrada/salida de empleados.
    
    POST /api/attendance/marcar/
    Body: {
        "tipo": "ENTRADA" | "SALIDA",
        "latitud": 4.6097,
        "longitud": -74.0817,
        "empleado_id": 1  (opcional, si no se envía marca para el usuario autenticado)
    }
    """
    permission_classes = [AllowAny]  # Cambiar a IsAuthenticated en producción

    def post(self, request, *args, **kwargs):
        try:
            data = request.data
            tipo = data.get('tipo', 'ENTRADA')
            latitud = data.get('latitud')
            longitud = data.get('longitud')
            empleado_id = data.get('empleado_id')

            # Obtener el empleado
            if empleado_id:
                try:
                    empleado = Empleado.objects.get(pk=empleado_id)
                except Empleado.DoesNotExist:
                    return Response({
                        'success': False,
                        'message': 'Empleado no encontrado.'
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                # Si no hay empleado_id, usar el primer empleado activo (para pruebas)
                empleado = Empleado.objects.filter(estado='activo').first()
                if not empleado:
                    return Response({
                        'success': False,
                        'message': 'No hay empleados disponibles.'
                    }, status=status.HTTP_400_BAD_REQUEST)

            # Verificar si ya marcó entrada/salida hoy
            hoy = timezone.now().date()
            registros_hoy = RegistroAsistencia.objects.filter(
                empleado=empleado,
                fecha_hora__date=hoy
            )

            if tipo == 'ENTRADA':
                if registros_hoy.filter(tipo='ENTRADA').exists():
                    return Response({
                        'success': False,
                        'message': 'Ya has marcado entrada hoy.'
                    }, status=status.HTTP_400_BAD_REQUEST)
            elif tipo == 'SALIDA':
                if not registros_hoy.filter(tipo='ENTRADA').exists():
                    return Response({
                        'success': False,
                        'message': 'Debes marcar entrada antes de marcar salida.'
                    }, status=status.HTTP_400_BAD_REQUEST)
                if registros_hoy.filter(tipo='SALIDA').exists():
                    return Response({
                        'success': False,
                        'message': 'Ya has marcado salida hoy.'
                    }, status=status.HTTP_400_BAD_REQUEST)

            # Calcular si es tardanza (para entrada)
            es_tardanza = False
            minutos_atraso = 0
            if tipo == 'ENTRADA':
                hora_limite = time(9, 0)  # 9:00 AM
                hora_actual = timezone.now().time()
                if hora_actual > hora_limite:
                    es_tardanza = True
                    tiempo_limite = datetime.combine(timezone.now().date(), hora_limite)
                    tiempo_actual = datetime.combine(timezone.now().date(), hora_actual)
                    diferencia = tiempo_actual - tiempo_limite
                    minutos_atraso = int(diferencia.total_seconds() / 60)

            # Crear registro
            registro = RegistroAsistencia.objects.create(
                empleado=empleado,
                tipo=tipo,
                latitud=latitud,
                longitud=longitud,
                es_tardanza=es_tardanza,
                minutos_atraso=minutos_atraso
            )

            return Response({
                'success': True,
                'message': f'Asistencia marcada correctamente: {tipo}',
                'registro': {
                    'id': registro.id,
                    'empleado': empleado.nombre_completo,
                    'tipo': registro.tipo,
                    'fecha_hora': registro.fecha_hora.strftime('%Y-%m-%d %H:%M:%S'),
                    'es_tardanza': registro.es_tardanza,
                    'minutos_atraso': registro.minutos_atraso
                }
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({
                'success': False,
                'message': f'Error interno del servidor: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AsistenciaHoyView(APIView):
    """
    API para obtener registros de asistencia del día actual.
    
    GET /api/attendance/today/
    Retorna lista de registros con coordenadas para mapa.
    """
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        hoy = timezone.now().date()
        registros = RegistroAsistencia.objects.filter(
            fecha_hora__date=hoy,
            tipo='ENTRADA'
        ).select_related('empleado')

        data = []
        for reg in registros:
            estado = 'A tiempo' if not reg.es_tardanza else 'Tarde'
            
            data.append({
                'id': reg.id,
                'empleado_nombre': reg.empleado.nombre_completo,
                'estado': estado,
                'lat': float(reg.latitud) if reg.latitud else None,
                'lng': float(reg.longitud) if reg.longitud else None,
                'fecha_hora': reg.fecha_hora.isoformat(),
                'minutos_atraso': reg.minutos_atraso
            })

        return Response(data)


class RegistroAsistenciaViewSet(viewsets.ModelViewSet):
    """
    API ViewSet para gestión de Registros de Asistencia.
    
    Endpoints:
    - GET /api/attendance/registros/     → Listar todos
    - POST /api/attendance/registros/    → Crear nuevo
    - GET /api/attendance/registros/{id}/    → Obtener uno
    - PUT/PATCH /api/attendance/registros/{id}/  → Actualizar
    - DELETE /api/attendance/registros/{id}/     → Eliminar
    """
    queryset = RegistroAsistencia.objects.select_related('empleado').all()
    serializer_class = RegistroAsistenciaSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['empleado', 'tipo', 'es_tardanza']
    ordering_fields = ['fecha_hora']
    ordering = ['-fecha_hora']

    def get_queryset(self):
        """Filtros adicionales por query params"""
        queryset = super().get_queryset()
        
        # Filtro por rango de fechas
        fecha_inicio = self.request.query_params.get('fecha_inicio', None)
        fecha_fin = self.request.query_params.get('fecha_fin', None)
        
        if fecha_inicio:
            queryset = queryset.filter(fecha_hora__date__gte=fecha_inicio)
        if fecha_fin:
            queryset = queryset.filter(fecha_hora__date__lte=fecha_fin)
        
        return queryset


class TurnoViewSet(viewsets.ModelViewSet):
    queryset = Turno.objects.select_related('empresa').all()
    serializer_class = TurnoSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['empresa']


class GeocercaViewSet(viewsets.ModelViewSet):
    queryset = Geocerca.objects.select_related('empresa').all()
    serializer_class = GeocercaSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['empresa', 'activo']


class ReglaAsistenciaViewSet(viewsets.ModelViewSet):
    queryset = ReglaAsistencia.objects.select_related('empresa', 'geocerca').all()
    serializer_class = ReglaAsistenciaSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['empresa', 'geocerca']


def _haversine_distance_m(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Distancia Haversine en metros."""
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
    c = 2 * asin(sqrt(a))
    r = 6371000
    return c * r


def _point_in_polygon(point_lat: float, point_lng: float, polygon: list) -> bool:
    """Algoritmo ray casting para polígono plano."""
    num = len(polygon)
    j = num - 1
    inside = False
    for i in range(num):
        lat_i, lng_i = polygon[i]['lat'], polygon[i]['lng']
        lat_j, lng_j = polygon[j]['lat'], polygon[j]['lng']
        intersect = ((lng_i > point_lng) != (lng_j > point_lng)) and (
            point_lat < (lat_j - lat_i) * (point_lng - lng_i) / (lng_j - lng_i + 1e-9) + lat_i
        )
        if intersect:
            inside = not inside
        j = i
    return inside


def _esta_dentro_geocerca(geocerca: Geocerca, lat: float, lng: float) -> bool:
    if not geocerca.activo or not geocerca.coordenadas:
        return False
    if geocerca.tipo == 'circulo':
        center = geocerca.coordenadas.get('center', {})
        radius = geocerca.coordenadas.get('radius_m', 0)
        if not center or not radius:
            return False
        distancia = _haversine_distance_m(lat, lng, center.get('lat'), center.get('lng'))
        return distancia <= radius
    if geocerca.tipo == 'poligono':
        puntos = geocerca.coordenadas if isinstance(geocerca.coordenadas, list) else []
        return _point_in_polygon(lat, lng, puntos)
    return False


class EventoAsistenciaViewSet(viewsets.ModelViewSet):
    queryset = EventoAsistencia.objects.select_related('empleado').all()
    serializer_class = EventoAsistenciaSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['empleado', 'tipo', 'fuente', 'dentro_geocerca']
    ordering_fields = ['registrado_el']
    ordering = ['-registrado_el']

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        empleado = serializer.validated_data['empleado']
        lat = serializer.validated_data.get('gps_lat')
        lng = serializer.validated_data.get('gps_lng')

        regla = ReglaAsistencia.objects.filter(empresa=empleado.empresa).select_related('geocerca').first()
        dentro_geocerca = False

        if regla and regla.geocerca:
            if lat is None or lng is None:
                return Response({'detail': 'Coordenadas requeridas para validar geocerca.'}, status=status.HTTP_400_BAD_REQUEST)
            dentro_geocerca = _esta_dentro_geocerca(regla.geocerca, float(lat), float(lng))
            if not dentro_geocerca:
                return Response({'detail': 'Marcación fuera de la geocerca permitida.'}, status=status.HTTP_400_BAD_REQUEST)

        evento = serializer.save(dentro_geocerca=dentro_geocerca, fuente=request.data.get('fuente', 'web'))
        headers = self.get_success_headers(serializer.data)
        return Response(self.get_serializer(evento).data, status=status.HTTP_201_CREATED, headers=headers)


class JornadaCalculadaViewSet(viewsets.ModelViewSet):
    queryset = JornadaCalculada.objects.select_related('empleado').all()
    serializer_class = JornadaCalculadaSerializer
    permission_classes = [AllowAny]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['empleado', 'fecha', 'estado']


class ExportarAsistenciaExcelView(APIView):
    """
    API para exportar pre-nómina en Excel.
    
    GET /api/attendance/exportar-excel/
    Retorna archivo Excel con datos de asistencia del mes actual.
    """
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Pre-Nómina"

        # Estilos
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Encabezados
        headers = [
            'Nombre Empleado', 
            'Días Trabajados', 
            'Horas Extra (Calculadas)', 
            'Minutos de Atraso',
            'Estado'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Obtener datos del mes actual
        hoy = timezone.now()
        inicio_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # Calcular fin de mes
        if hoy.month == 12:
            fin_mes = hoy.replace(year=hoy.year + 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        else:
            fin_mes = hoy.replace(month=hoy.month + 1, day=1, hour=0, minute=0, second=0, microsecond=0)

        # Obtener empleados activos
        empleados = Empleado.objects.filter(estado='activo')

        row_num = 2
        for empleado in empleados:
            # Calcular días trabajados
            dias_trabajados = RegistroAsistencia.objects.filter(
                empleado=empleado,
                tipo='ENTRADA',
                fecha_hora__gte=inicio_mes,
                fecha_hora__lt=fin_mes
            ).count()

            # Calcular horas extra (simplificado)
            horas_extra = 0
            registros_mes = RegistroAsistencia.objects.filter(
                empleado=empleado,
                fecha_hora__gte=inicio_mes,
                fecha_hora__lt=fin_mes
            ).order_by('fecha_hora')

            entrada_actual = None
            for registro in registros_mes:
                if registro.tipo == 'ENTRADA':
                    entrada_actual = registro.fecha_hora
                elif registro.tipo == 'SALIDA' and entrada_actual and entrada_actual.date() == registro.fecha_hora.date():
                    horas_trabajadas = (registro.fecha_hora - entrada_actual).total_seconds() / 3600
                    if horas_trabajadas > 8:
                        horas_extra += horas_trabajadas - 8
                    entrada_actual = None

            # Calcular minutos de atraso
            minutos_atraso = RegistroAsistencia.objects.filter(
                empleado=empleado,
                tipo='ENTRADA',
                es_tardanza=True,
                fecha_hora__gte=inicio_mes,
                fecha_hora__lt=fin_mes
            ).aggregate(total=Sum('minutos_atraso'))['total'] or 0

            # Estado
            if minutos_atraso == 0:
                estado = 'Excelente'
            elif minutos_atraso < 30:
                estado = 'Bueno'
            else:
                estado = 'Requiere Atención'

            # Escribir fila
            ws.cell(row=row_num, column=1, value=str(empleado))
            ws.cell(row=row_num, column=2, value=dias_trabajados)
            ws.cell(row=row_num, column=3, value=round(horas_extra, 2))
            ws.cell(row=row_num, column=4, value=minutos_atraso)
            ws.cell(row=row_num, column=5, value=estado)

            row_num += 1

        # Ajustar ancho de columnas
        for col in range(1, 6):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 25

        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=pre-nomina_{hoy.strftime("%Y-%m")}.xlsx'

        # Guardar workbook en la respuesta
        wb.save(response)

        return response
